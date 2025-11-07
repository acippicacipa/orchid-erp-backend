from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from .models import ImportTemplate, DataImport, ImportLog
from .services import DataImportService, TemplateService
from accounts.permissions import CanImportData, CanViewImportHistory, CanDownloadTemplates
import openpyxl
from io import BytesIO

class DataUploadView(APIView):
    """
    Handle file upload for data import
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        file = request.FILES.get('file')
        template_id = request.data.get('template_id')
        
        if not file:
            return Response({
                'error': 'No file provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not template_id:
            return Response({
                'error': 'Template ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get template
        try:
            template = ImportTemplate.objects.get(id=template_id, is_active=True)
        except ImportTemplate.DoesNotExist:
            return Response({
                'error': 'Template not found'
            }, status=status.HTTP_404_NOT_FOUND)

        try:
            # Buat record DataImport dengan relasi ke template
            data_import = DataImport.objects.create(
                template=template,
                file=file,
                original_filename=file.name,
                uploaded_by=request.user,
                status='PENDING'
            )
        
            # Validate file using service
            service = DataImportService(data_import.id)
            service.process_file()
            data_import.refresh_from_db()
            
            return Response({
                'import_id': data_import.id,
                'status': data_import.status,
                'total_rows': data_import.total_rows,
                'successful_rows': data_import.successful_rows,
                'failed_rows': data_import.failed_rows,
                'message': 'File processed successfully.'
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            # Log error untuk debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Upload failed: {e}", exc_info=True)
            return Response({'error': f'An unexpected error occurred: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TemplateDownloadView(APIView):
    """
    Download import templates
    """
    permission_classes = [IsAuthenticated, CanDownloadTemplates]
    
    def get(self, request):
        try:
            template_id = request.query_params.get('template_id')
            
            if not template_id:
                # Return list of available templates filtered by user permissions
                templates = ImportTemplate.objects.filter(is_active=True)
                template_data = []
                
                for template in templates:
                    # Check if user can access this template type
                    if self._can_access_template(request.user, template.template_type):
                        template_data.append({
                            'id': template.id,
                            'name': template.name,
                            'template_type': template.template_type,
                            'description': template.description,
                            'required_columns': template.required_columns,
                            'optional_columns': template.optional_columns
                        })
                
                return Response({
                    'templates': template_data
                }, status=status.HTTP_200_OK)
            
            # Download specific template
            try:
                template = ImportTemplate.objects.get(id=template_id, is_active=True)
            except ImportTemplate.DoesNotExist:
                return Response({
                    'error': 'Template not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Create Excel template file
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = template.name
            
            # Add headers
            all_columns = template.required_columns + template.optional_columns
            for col_idx, column in enumerate(all_columns, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column)
                # Mark required columns in bold
                if column in template.required_columns:
                    cell.font = openpyxl.styles.Font(bold=True)
            
            # Add sample data row based on template type
            sample_data = {}
            if template.template_type == 'MAIN_CATEGORIES':
                sample_data = {
                    'name': 'Barang Lokal',
                    'description': 'Kategori untuk barang yang diproduksi di dalam negeri',
                    'is_active': True
                }
            elif template.template_type == 'SUB_CATEGORIES':
                sample_data = {
                    'name': 'Bunga',
                    'description': 'Sub-kategori untuk produk bunga',
                    'is_active': True
                }
            elif template.template_type == 'CATEGORIES':
                sample_data = {
                    'name': 'Bunga Lokal',
                    'main_category': 'Barang Lokal',
                    'sub_category': 'Bunga',
                    'description': 'Kategori gabungan Bunga Lokal',
                    'is_active': True
                }
            elif template.template_type == 'LOCATIONS':
                sample_data = {
                    'name': 'Gudang Utama',
                    'code': 'GU001',
                    'location_type': 'WAREHOUSE',
                    'address': 'Jl. Merdeka No. 1',
                    'contact_person': 'Budi',
                    'phone': '08123456789',
                    'email': 'gudang@example.com',
                    'is_active': True,
                    'is_sellable_location': True,
                    'is_purchasable_location': True,
                    'is_manufacturing_location': False,
                    'storage_capacity': 1000.00,
                    'current_utilization': 0.00,
                    'notes': 'Gudang utama untuk semua stok'
                }
            elif template.template_type == 'ITEMS':
                sample_data = {
                    'name': 'Mawar Merah',
                    'sku': 'MR-001',
                    'description': 'Bunga Mawar Merah Lokal',
                    'main_category': 'Barang Lokal',
                    'sub_category': 'Bunga',
                    'color': 'Merah',
                    'size': 'Besar',
                    'brand': 'Flora Nusantara',
                    'model': 'Mawar',
                    'cost_price': 10000.00,
                    'selling_price': 15000.00,
                    'discount': 0.00,
                    'unit_of_measure': 'tangkai',
                    'weight': 0.1,
                    'dimensions': '10x5x50',
                    'is_active': True,
                    'is_sellable': True,
                    'is_purchasable': True,
                    'is_manufactured': False,
                    'minimum_stock_level': 10,
                    'maximum_stock_level': 100,
                    'reorder_point': 20,
                    'barcode': '1234567890123',
                    'supplier_code': 'SUP-MR-001',
                    'notes': 'Produk unggulan'
                }
            elif template.template_type == 'INVENTORY':
                sample_data = {
                    'product_sku': 'MR-001',
                    'location_code': 'GU001',
                    'quantity_on_hand': 50.00,
                    'quantity_sellable': 50.00,
                    'quantity_non_sellable': 0.00,
                    'quantity_reserved': 0.00,
                    'quantity_allocated': 0.00,
                    'average_cost': 10000.00,
                    'last_cost': 10000.00,
                    'bin_location': 'A-01-01',
                    'lot_number': 'LOT-20231027',
                    'expiry_date': '2024-12-31',
                    'notes': 'Stok awal'
                }
            elif template.template_type == 'CUSTOMERS':
                sample_data = {
                    'name': 'Acme Corp',
                    'customer_id': 'CUST001',
                    'email': 'contact@acmecorp.com',
                    'phone': '123-456-7890',
                    'contact_person': 'John Doe',
                    'address_line_1': '123 Main St',
                    'city': 'Anytown',
                    'state': 'CA',
                    'postal_code': '90210',
                    'country': 'USA'
                }
            elif template.template_type == 'SUPPLIERS':
                sample_data = {
                    'name': 'Global Supplies Inc',
                    'supplier_id': 'SUP001',
                    'email': 'info@globalsupplies.com',
                    'phone': '987-654-3210',
                    'contact_person': 'Jane Smith',
                    'address_line_1': '456 Oak Ave',
                    'city': 'Otherville',
                    'state': 'NY',
                    'postal_code': '10001',
                    'country': 'USA'
                }
            elif template.template_type == 'ITEMS':
                sample_data = {
                    'name': 'Laptop Pro X',
                    'sku': 'LPX-001',
                    'description': 'High performance laptop',
                    'price': 1200.00,
                    'category_name': 'Electronics',
                    'category_code': 'ELEC'
                }
            elif template.template_type == 'USERS':
                sample_data = {
                    'username': 'newuser',
                    'email': 'newuser@example.com',
                    'password': 'password123',
                    'first_name': 'New',
                    'last_name': 'User',
                    'role': 'SALES',
                    'employee_id': 'EMP005',
                    'department': 'Sales',
                    'position': 'Sales Rep',
                    'hire_date': '2023-01-15',
                    'is_active': True
                }
            elif template.template_type == 'SALES_ORDERS':
                sample_data = {
                    'customer_id': 'CUST001',
                    'order_date': '2025-09-18',
                    'due_date': '2025-10-18',
                    'order_number': 'SO-2025-001',
                    'status': 'CONFIRMED',
                    'total_amount': 1500.00,
                    'discount_amount': 50.00,
                    'tax_amount': 100.00,
                    'notes': 'Urgent order',
                    'shipping_address_line_1': '123 Main St',
                    'shipping_city': 'Anytown',
                    'shipping_state': 'CA',
                    'shipping_postal_code': '90210',
                    'billing_address_line_1': '123 Main St',
                    'billing_city': 'Anytown',
                    'billing_state': 'CA',
                    'billing_postal_code': '90210',
                }
            elif template.template_type == 'INVOICES':
                sample_data = {
                    'customer_id': 'CUST001',
                    'invoice_date': '2025-09-18',
                    'due_date': '2025-10-18',
                    'invoice_number': 'INV-2025-001',
                    'sales_order_number': 'SO-2025-001',
                    'status': 'SENT',
                    'total_amount': 1500.00,
                    'amount_paid': 0.00,
                    'balance_due': 1500.00,
                    'notes': 'Initial invoice'
                }
            elif template.template_type == 'PAYMENTS':
                sample_data = {
                    'invoice_number': 'INV-2025-001',
                    'payment_date': '2025-09-18',
                    'amount': 500.00,
                    'payment_method': 'BANK_TRANSFER',
                    'transaction_id': 'TXN12345',
                    'notes': 'Partial payment'
                }
            elif template.template_type == 'PURCHASE_ORDERS':
                sample_data = {
                    'supplier_id': 'SUP001',
                    'order_date': '2025-09-18',
                    'expected_delivery_date': '2025-10-18',
                    'order_number': 'PO-2025-001',
                    'status': 'CONFIRMED',
                    'total_amount': 800.00,
                    'notes': 'Office supplies'
                }
            elif template.template_type == 'BILLS':
                sample_data = {
                    'supplier_id': 'SUP001',
                    'bill_date': '2025-09-18',
                    'due_date': '2025-10-18',
                    'bill_number': 'BILL-2025-001',
                    'purchase_order_number': 'PO-2025-001',
                    'status': 'PENDING',
                    'total_amount': 800.00,
                    'amount_paid': 0.00,
                    'balance_due': 800.00,
                    'notes': 'Monthly supplies bill'
                }
            elif template.template_type == 'SUPPLIER_PAYMENTS':
                sample_data = {
                    'bill_number': 'BILL-2025-001',
                    'payment_date': '2025-09-18',
                    'amount': 400.00,
                    'payment_method': 'BANK_TRANSFER',
                    'transaction_id': 'TXN67890',
                    'notes': 'Partial payment for bill'
                }
            elif template.template_type == 'ACCOUNTS':
                sample_data = {
                    'name': 'Cash Account',
                    'account_type': 'ASSET',
                    'account_number': '1000',
                    'description': 'Main operating cash account',
                    'is_active': True
                }
            elif template.template_type == 'JOURNAL_ENTRIES':
                sample_data = {
                    'entry_date': '2025-09-18',
                    'reference_number': 'JE-2025-001',
                    'entry_type': 'GENERAL',
                    'description': 'Recording office supplies purchase',
                    'total_debit': 100.00,
                    'total_credit': 100.00,
                    'is_posted': False
                }
            elif template.template_type == 'JOURNAL_ITEMS':
                sample_data = {
                    'journal_entry_reference_number': 'JE-2025-001',
                    'account_number': '5000',
                    'debit': 100.00,
                    'credit': 0.00,
                    'description': 'Office Supplies Expense'
                }
            elif template.template_type == 'LEDGERS':
                sample_data = {
                    'account_number': '1000',
                    'journal_item_id': 1, # This would need to be a valid ID from an existing JournalItem
                    'date': '2025-09-18',
                    'description': 'Initial cash entry',
                    'debit': 10000.00,
                    'credit': 0.00,
                    'balance': 10000.00
                }

            for col_idx, column in enumerate(all_columns, 1):
                if column in sample_data:
                    worksheet.cell(row=2, column=col_idx, value=sample_data[column])
            
            # Save to BytesIO
            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)
            
            # Create response
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{template.name}_template.xlsx"'
            
            return response
            
        except Exception as e:
            return Response({
                'error': f'Template download failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _can_access_template(self, user, template_type):
        """Check if user can access specific template type"""
        # Admin can access all templates
        if user.profile.has_role("ADMIN"):
            return True
        
        # Sales module templates
        if template_type in ['CUSTOMERS', 'SALES_ORDERS', 'INVOICES', 'PAYMENTS']:
            return user.profile.has_role("SALES")
        
        # Inventory module templates
        if template_type in ['ITEMS', 'CATEGORIES', 'INVENTORY']:
            return user.profile.has_role("WAREHOUSE")
        
        # Purchasing module templates
        if template_type in ['SUPPLIERS', 'PURCHASE_ORDERS', 'BILLS', 'SUPPLIER_PAYMENTS']:
            return user.profile.has_role("PURCHASING")
        
        # Accounting module templates
        if template_type in ['ACCOUNTS', 'JOURNAL_ENTRIES', 'JOURNAL_ITEMS', 'LEDGERS']:
            return user.profile.has_role("ACCOUNTING")
        
        # Common templates (locations, users) - only admin
        if template_type in ['LOCATIONS', 'USERS']:
            return user.profile.has_role("ADMIN")
        
        return False

class ImportHistoryView(APIView):
    """
    Get import history for the current user
    """
    permission_classes = [IsAuthenticated, CanViewImportHistory]
    
    def get(self, request):
        try:
            # Filter imports based on user permissions
            if request.user.profile.has_role("ADMIN") or request.user.profile.has_role("AUDIT"):
                # Admin and Audit can see all imports
                imports = DataImport.objects.all().order_by('-created_at')
            else:
                # Other users can only see their own imports
                imports = DataImport.objects.filter(created_by=request.user).order_by('-created_at')
            
            import_data = []
            for data_import in imports:
                success_rate = 0
                if data_import.total_rows > 0:
                    # Gunakan successful_rows untuk menghitung rate
                    success_rate = round((data_import.successful_rows / data_import.total_rows) * 100, 2)

                import_data.append({
                    'id': data_import.id,
                    'template_name': data_import.template.name,
                    'template_type': data_import.template.template_type,
                    'original_filename': data_import.original_filename,
                    'status': data_import.status,
                    'total_rows': data_import.total_rows,
                    
                    # Ganti field yang salah dengan yang benar
                    'successful_rows': data_import.successful_rows, # Ganti 'valid_rows' dan 'imported_rows'
                    'failed_rows': data_import.failed_rows,       # Ganti 'invalid_rows'
                    'processed_rows': data_import.processed_rows,
                    
                    # Gunakan success_rate yang baru dihitung
                    'success_rate': success_rate,
                    
                    'created_at': data_import.created_at,
                    'started_at': data_import.started_at,
                    'completed_at': data_import.completed_at,
                    'uploaded_by': data_import.uploaded_by.username if data_import.uploaded_by else None
                })
            
            return Response({
                'imports': import_data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Failed to get import history: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def delete(self, request, import_id):
        """
        Menangani permintaan DELETE untuk menghapus satu record DataImport.
        """
        try:
            # Cari record import yang akan dihapus.
            # Filter juga berdasarkan pengguna jika Anda tidak ingin admin bisa menghapus milik semua orang.
            # import_to_delete = DataImport.objects.get(id=import_id, uploaded_by=request.user)
            import_to_delete = DataImport.objects.get(id=import_id)
            
            # Hapus objek dari database.
            # File fisik yang terkait (di folder media) juga akan terhapus jika
            # Anda menggunakan Django-Cleanup atau menangani sinyal post_delete.
            # Untuk saat ini, kita hanya hapus record database.
            import_to_delete.delete()
            
            # Kirim respons sukses tanpa konten.
            return Response(status=status.HTTP_204_NO_CONTENT)

        except DataImport.DoesNotExist:
            return Response({'error': 'Import record not found.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to delete import record {import_id}: {e}", exc_info=True)
            return Response({'error': f'An unexpected error occurred: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ImportLogsView(APIView):
    """
    Get detailed logs for a specific import
    """
    permission_classes = [IsAuthenticated, CanViewImportHistory]
    
    def get(self, request):
        try:
            data_import_id = request.query_params.get('import_id')
            
            if not data_import_id:
                return Response({
                    'error': 'Import ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                data_import = DataImport.objects.get(id=data_import_id, created_by=request.user)
            except DataImport.DoesNotExist:
                return Response({
                    'error': 'Data import not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get logs
            logs = ImportLog.objects.filter(data_import=data_import).order_by('timestamp')
            
            log_data = []
            for log in logs:
                log_data.append({
                    'timestamp': log.timestamp,
                    'level': log.level,
                    'message': log.message,
                    'details': log.details
                })
            
            return Response({
                'import_id': data_import.id,
                'logs': log_data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Failed to get import logs: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SetupTemplatesView(APIView):
    """
    Setup default import templates
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # Check if user has admin permissions
            if not request.user.is_staff and not request.user.is_superuser:
                return Response({
                    'error': 'Admin permissions required'
                }, status=status.HTTP_403_FORBIDDEN)
            
            TemplateService.create_default_templates()
            
            return Response({
                'message': 'Default templates created successfully'
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Failed to create templates: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

